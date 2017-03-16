#!/usr/bin/env python
# _*_ coding: utf-8 _*_

import sys
import datetime
import openpyxl

def writetest(target, iso_dates=True):
    dd = range(0, 3) + range(57, 61)
    dates = [datetime.date(1900, 1, 1) + datetime.timedelta(days=d) for d in dd]
    isodates = [d.isoformat() for d in dates]
    tt = [(h, m, s) for h, m, s in zip(range(0, 4) + range(21, 24),
                                       range(0, 4) + range(57, 60),
                                       range(0, 4) + range(57, 60))]
    times = [datetime.time(h, m, s) for h, m, s in  tt]
    isotimes = [t.isoformat() for t in times]
    datetimes = [datetime.datetime(d.year, d.month, d.day,
                                   t.hour, t.minute, t.second)
                 for d, t in zip(dates, times)]

    if iso_dates:
        wb = openpyxl.Workbook(iso_dates=True)
        print('New workbook opened with iso_dates=True')
    else:
        wb = openpyxl.Workbook()
        print('New workbook opened without iso_dates argument')

    ws = wb.active
    ws.append(['string_date', 'written_date',
               'string_time', 'written_time', 'written_datetime',
               'read_date', 'read_time', 'read_datetime'])

    for d in zip(isodates, dates, isotimes, times, datetimes):
        ws.append(d)
    wb.save(target)

def readwritetest(source, target):
    wb = openpyxl.load_workbook(source)
    ws = wb.active
    dates = [d[0].value for d in ws['B2':'B8']]
    times = [d[0].value for d in ws['D2':'D8']]
    datetimes = [d[0].value for d in ws['E2':'E8']]
    for i, row in enumerate(range(2, 9)):
        ws.cell(row=row, column=6).value = dates[i]
        ws.cell(row=row, column=7).value = times[i]
        ws.cell(row=row, column=8).value = datetimes[i]
    wb.save(target)

if __name__ == '__main__':
    # Roundtrip
    writetest(target='xlsx/testdates_firstwrite.xlsx', iso_dates=False)
    readwritetest(source='xlsx/testdates_firstwrite.xlsx',
                  target='xlsx/testdates_round.xlsx')
    
    writetest(target='xlsx/testdates_firstwrite_isodates.xlsx', iso_dates=True)
    readwritetest(source='xlsx/testdates_firstwrite_isodates.xlsx',
                  target='xlsx/testdates_round_isodates.xlsx')
    
    ## Excel Interrupt
    print("Open {} in Excel and immediatel save it, unchanged, as {}, close Excel"
          .format('testdates_firstwrite.xlsx', 'testdates_firstwrite_copy.xlsx'))
    raw_input("hit any key when done")
    readwritetest(source='xlsx/testdates_firstwrite_copy.xlsx',
                  target='xlsx/testdates_xinterrupt.xlsx')
    
    print("Open {} in Excel and immediatel save it, unchanged, as {}, close Excel"
          .format('testdates_firstwrite_isodates.xlsx',
                  'testdates_firstwrite_isodates_copy.xlsx'))
    raw_input("hit any key when done")
    readwritetest(source='xlsx/testdates_firstwrite_isodates_copy.xlsx',
                  target='xlsx/testdates_isodates_xinterrupt.xlsx')
     
    
